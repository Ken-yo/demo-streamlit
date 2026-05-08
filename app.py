from __future__ import annotations

import os
import sqlite3
import threading
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

APP_TITLE = "社内コース入力・集計システム"
COURSE_COLUMNS = ["Tableau", "RPA", "ビジネスコア", "DBエンジニア", "プロ"]
COURSE_TO_DB = {"Tableau": "tableau", "RPA": "rpa", "ビジネスコア": "business_core", "DBエンジニア": "db_engineer", "プロ": "pro"}
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = Path(os.getenv("DATA_DIR", BASE_DIR / "data"))
DB_PATH = Path(os.getenv("DB_PATH", DATA_DIR / "course_entries.db"))
EXCEL_PATH = Path(os.getenv("EXCEL_PATH", DATA_DIR / "course_entries.xlsx"))
EXPORT_LOCK = threading.Lock()
DEPARTMENT_OPTIONS = ["(HB企)", "(HB技)", "(MR)", "(HB制振)", "(FR)", "(生活インフラ)", "(加)"]
FIXED_INPUT_ROWS = [
    {"年": 2026, "時期": "上期"}, {"年": 2026, "時期": "下期"},
    {"年": 2027, "時期": "上期"}, {"年": 2027, "時期": "下期"},
    {"年": 2028, "時期": "上期"}, {"年": 2028, "時期": "下期"},
    {"年": 2029, "時期": "上期"}, {"年": 2029, "時期": "下期"},
    {"年": 2030, "時期": "上期"}, {"年": 2030, "時期": "下期"},
]
SYSTEM_EMPLOYEE_NAME = ""


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except TypeError:
        pass
    return str(value).strip()


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except TypeError:
        pass
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None
        value = value.replace(",", "")
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if number.is_integer():
        return int(number)
    return number


def format_number_for_display(value: float | int) -> float | int:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return value
    if number.is_integer():
        return int(number)
    return number


def convert_course_columns_to_numeric(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    df = df.copy()
    for course in COURSE_COLUMNS:
        if course in df.columns:
            df[course] = pd.to_numeric(df[course], errors="coerce").fillna(0)
            df[course] = df[course].map(format_number_for_display)
    return df


def make_input_total_df(input_df: pd.DataFrame) -> pd.DataFrame:
    total_row: dict[str, Any] = {"年": "", "時期": "合計"}
    for course in COURSE_COLUMNS:
        total = pd.to_numeric(input_df.get(course, pd.Series(dtype=float)), errors="coerce").fillna(0).sum()
        total_row[course] = format_number_for_display(total)
    return pd.DataFrame([total_row], columns=["年", "時期", *COURSE_COLUMNS])


def get_conn() -> sqlite3.Connection:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH, timeout=30, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=10000")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def ensure_column_exists(conn: sqlite3.Connection, table_name: str, column_name: str, column_definition: str) -> None:
    existing_columns = {row["name"] for row in conn.execute(f"PRAGMA table_info({table_name})").fetchall()}
    if column_name not in existing_columns:
        conn.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_definition}")


def init_db() -> None:
    with get_conn() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS course_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                department TEXT NOT NULL,
                employee_name TEXT NOT NULL,
                target_year INTEGER NOT NULL,
                period TEXT NOT NULL,
                record_label TEXT DEFAULT '',
                tableau REAL DEFAULT 0,
                rpa REAL DEFAULT 0,
                business_core REAL DEFAULT 0,
                db_engineer REAL DEFAULT 0,
                pro REAL DEFAULT 0,
                memo TEXT DEFAULT '',
                created_by TEXT DEFAULT '',
                created_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_course_records_person_period
            ON course_records(department, employee_name, target_year, period)
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS deleted_course_records (
                deleted_id INTEGER PRIMARY KEY AUTOINCREMENT,
                original_id INTEGER NOT NULL,
                department TEXT NOT NULL,
                employee_name TEXT NOT NULL,
                target_year INTEGER NOT NULL,
                period TEXT NOT NULL,
                record_label TEXT DEFAULT '',
                tableau REAL DEFAULT 0,
                rpa REAL DEFAULT 0,
                business_core REAL DEFAULT 0,
                db_engineer REAL DEFAULT 0,
                pro REAL DEFAULT 0,
                memo TEXT DEFAULT '',
                created_by TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                deleted_by TEXT DEFAULT '',
                delete_comment TEXT NOT NULL,
                deleted_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_deleted_course_records_original_id
            ON deleted_course_records(original_id)
        """)
        ensure_column_exists(conn, "course_records", "business_core", "REAL DEFAULT 0")
        ensure_column_exists(conn, "deleted_course_records", "business_core", "REAL DEFAULT 0")
        conn.execute("""
            CREATE TABLE IF NOT EXISTS audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                action TEXT NOT NULL,
                record_id INTEGER,
                department TEXT,
                employee_name TEXT,
                actor TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY(record_id) REFERENCES course_records(id)
            )
        """)


def make_default_input_df() -> pd.DataFrame:
    rows = []
    for fixed in FIXED_INPUT_ROWS:
        rows.append({
            "年": fixed["年"], "時期": fixed["時期"],
            "Tableau": 0, "RPA": 0, "ビジネスコア": 0, "DBエンジニア": 0, "プロ": 0,
        })
    return pd.DataFrame(rows)


def normalize_input_rows(input_df: pd.DataFrame, comment: str) -> tuple[list[dict[str, Any]], list[str]]:
    errors: list[str] = []
    rows: list[dict[str, Any]] = []
    comment = clean_text(comment)
    if not comment:
        errors.append("コメントを入力してください。")
    if len(input_df) != len(FIXED_INPUT_ROWS):
        errors.append(f"入力テーブルは{len(FIXED_INPUT_ROWS)}行で登録してください。")

    for idx, fixed in enumerate(FIXED_INPUT_ROWS):
        if idx >= len(input_df):
            errors.append(f"{idx + 1}行目がありません。")
            continue
        row = input_df.iloc[idx]
        target_year = int(fixed["年"])
        period = str(fixed["時期"])
        values: dict[str, float | int] = {}
        for course, db_col in COURSE_TO_DB.items():
            number = parse_number(row.get(course, None))
            if number is None:
                errors.append(f"{idx + 1}行目（{target_year} {period}）の{course}を数値で入力してください。")
                continue
            values[db_col] = number
        rows.append({"target_year": target_year, "period": period, "record_label": f"{target_year}_{period}", **values, "memo": comment})
    if errors:
        return [], errors
    return rows, []


def insert_records(department: str, comment: str, rows: list[dict[str, Any]]) -> int:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    employee_name = SYSTEM_EMPLOYEE_NAME
    actor = department
    values = [(
        department, employee_name, int(row["target_year"]), clean_text(row["period"]), row.get("record_label", ""),
        row.get("tableau", 0), row.get("rpa", 0), row.get("business_core", 0), row.get("db_engineer", 0), row.get("pro", 0),
        comment, actor, now,
    ) for row in rows]
    if not values:
        return 0
    with get_conn() as conn:
        conn.execute("BEGIN IMMEDIATE")
        conn.executemany("""
            INSERT INTO course_records (
                department, employee_name, target_year, period, record_label,
                tableau, rpa, business_core, db_engineer, pro, memo, created_by, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, values)
        inserted_ids = [row["id"] for row in conn.execute(
            "SELECT id FROM course_records ORDER BY id DESC LIMIT ?", (len(values),)
        ).fetchall()]
        for record_id in inserted_ids:
            conn.execute("""
                INSERT INTO audit_logs(action, record_id, department, employee_name, actor, created_at)
                VALUES (?, ?, ?, ?, ?, ?)
            """, ("insert", record_id, department, employee_name, actor, now))
        conn.commit()
    return len(values)


def delete_records(record_ids: list[int], delete_comment: str, deleted_by: str = "管理者") -> int:
    if not record_ids:
        return 0
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    placeholders = ",".join("?" for _ in record_ids)
    with get_conn() as conn:
        conn.execute("BEGIN IMMEDIATE")
        rows = conn.execute(f"""
            SELECT id, department, employee_name, target_year, period, record_label,
                   tableau, rpa, business_core, db_engineer, pro, memo, created_by, created_at
            FROM course_records
            WHERE id IN ({placeholders})
            ORDER BY id
        """, record_ids).fetchall()
        for row in rows:
            conn.execute("""
                INSERT INTO deleted_course_records (
                    original_id, department, employee_name, target_year, period, record_label,
                    tableau, rpa, business_core, db_engineer, pro, memo, created_by, created_at,
                    deleted_by, delete_comment, deleted_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                row["id"], row["department"], row["employee_name"], row["target_year"], row["period"], row["record_label"],
                row["tableau"], row["rpa"], row["business_core"], row["db_engineer"], row["pro"], row["memo"], row["created_by"], row["created_at"],
                deleted_by, delete_comment, now,
            ))
            conn.execute("""
                INSERT INTO audit_logs(action, record_id, department, employee_name, actor, created_at)
                VALUES (?, ?, ?, ?, ?, ?)
            """, ("delete", None, row["department"], row["employee_name"], deleted_by, now))
        conn.execute(f"UPDATE audit_logs SET record_id = NULL WHERE record_id IN ({placeholders})", record_ids)
        conn.execute(f"DELETE FROM course_records WHERE id IN ({placeholders})", record_ids)
        conn.commit()
    return len(rows)


def fetch_records() -> pd.DataFrame:
    columns = ["ID", "部署", "年", "時期", "Tableau", "RPA", "ビジネスコア", "DBエンジニア", "プロ", "コメント", "登録者", "登録日時"]
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT id, department, target_year, period, tableau, rpa, business_core, db_engineer, pro, memo, created_by, created_at
            FROM course_records
            ORDER BY target_year ASC, CASE period WHEN '上期' THEN 1 WHEN '下期' THEN 2 ELSE 9 END ASC, id ASC
        """).fetchall()
    if not rows:
        return pd.DataFrame(columns=columns)
    df = pd.DataFrame([dict(row) for row in rows]).rename(columns={
        "id": "ID", "department": "部署", "target_year": "年", "period": "時期",
        "tableau": "Tableau", "rpa": "RPA", "business_core": "ビジネスコア", "db_engineer": "DBエンジニア", "pro": "プロ",
        "memo": "コメント", "created_by": "登録者", "created_at": "登録日時",
    })
    df = df[columns]
    return convert_course_columns_to_numeric(df)


def fetch_detail_records() -> pd.DataFrame:
    aggregate_df = fetch_records()
    detail_rows: list[dict[str, Any]] = []
    for _, row in aggregate_df.iterrows():
        for course in COURSE_COLUMNS:
            value = parse_number(row.get(course, None))
            if value is not None:
                detail_rows.append({
                    "ID": row["ID"], "部署": row["部署"], "年": row["年"], "時期": row["時期"],
                    "コース": course, "入力された情報": value, "コメント": row["コメント"], "登録日時": row["登録日時"],
                })
    return pd.DataFrame(detail_rows, columns=["ID", "部署", "年", "時期", "コース", "入力された情報", "コメント", "登録日時"])


def fetch_deleted_records() -> pd.DataFrame:
    columns = [
        "削除履歴ID", "元ID", "部署", "年", "時期", "Tableau", "RPA", "ビジネスコア", "DBエンジニア", "プロ",
        "登録時コメント", "登録者", "登録日時", "削除者", "削除コメント", "削除日時",
    ]
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT deleted_id, original_id, department, target_year, period, tableau, rpa, business_core, db_engineer, pro,
                   memo, created_by, created_at, deleted_by, delete_comment, deleted_at
            FROM deleted_course_records
            ORDER BY deleted_at DESC, deleted_id DESC
        """).fetchall()
    if not rows:
        return pd.DataFrame(columns=columns)
    df = pd.DataFrame([dict(row) for row in rows]).rename(columns={
        "deleted_id": "削除履歴ID", "original_id": "元ID", "department": "部署", "target_year": "年", "period": "時期",
        "tableau": "Tableau", "rpa": "RPA", "business_core": "ビジネスコア", "db_engineer": "DBエンジニア", "pro": "プロ", "memo": "登録時コメント",
        "created_by": "登録者", "created_at": "登録日時", "deleted_by": "削除者", "delete_comment": "削除コメント", "deleted_at": "削除日時",
    })
    df = df[columns]
    return convert_course_columns_to_numeric(df)


def write_dataframe(ws, df: pd.DataFrame) -> None:
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))


def style_sheet_as_table(ws, table_name: str) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    max_row = max(ws.max_row, 1)
    max_col = max(ws.max_column, 1)
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    widths = {"A": 12, "B": 12, "C": 18, "D": 8, "E": 10, "F": 26, "G": 26, "H": 18, "I": 28, "J": 26, "K": 30, "L": 16, "M": 20, "N": 16, "O": 30, "P": 20}
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = widths.get(col_letter, 18)
    if max_row >= 2:
        ref = f"A1:{get_column_letter(max_col)}{max_row}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)
    else:
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}1"


def write_summary(ws, aggregate_df: pd.DataFrame, detail_df: pd.DataFrame, deleted_df: pd.DataFrame) -> None:
    ws["A1"] = "サマリー"
    ws["A1"].font = Font(bold=True, size=16)
    summary_items = [
        ("登録中件数", int(len(aggregate_df))),
        ("コース別明細件数", int(len(detail_df))),
        ("削除履歴件数", int(len(deleted_df))),
        ("最終Excel更新日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for idx, (label, value) in enumerate(summary_items, start=3):
        ws.cell(row=idx, column=1, value=label)
        ws.cell(row=idx, column=2, value=value)
    ws["A8"] = "コース"
    ws["B8"] = "入力あり件数"
    for row_no, course in enumerate(COURSE_COLUMNS, start=9):
        count = 0 if aggregate_df.empty else aggregate_df[course].fillna("").astype(str).str.strip().ne("").sum()
        ws.cell(row=row_no, column=1, value=course)
        ws.cell(row=row_no, column=2, value=int(count))
    ws["D8"] = "部署"
    ws["E8"] = "登録件数"
    if not aggregate_df.empty:
        for dept_row, (_, row) in enumerate(aggregate_df.groupby("部署", dropna=False).size().reset_index(name="登録件数").iterrows(), start=9):
            ws.cell(row=dept_row, column=4, value=row["部署"])
            ws.cell(row=dept_row, column=5, value=int(row["登録件数"]))
    for cell in ws[8]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows(min_row=1, max_row=max(ws.max_row, 11), max_col=5):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col, width in {"A": 20, "B": 18, "D": 24, "E": 14}.items():
        ws.column_dimensions[col].width = width


def sync_excel_from_sqlite() -> Path:
    with EXPORT_LOCK:
        EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
        aggregate_df = fetch_records()
        detail_df = fetch_detail_records()
        deleted_df = fetch_deleted_records()
        wb = Workbook()
        ws = wb.active
        ws.title = "集計テーブル"
        write_dataframe(ws, aggregate_df)
        style_sheet_as_table(ws, "AggregateTable")
        detail_ws = wb.create_sheet("コース別明細")
        write_dataframe(detail_ws, detail_df)
        style_sheet_as_table(detail_ws, "DetailTable")
        deleted_ws = wb.create_sheet("削除履歴")
        write_dataframe(deleted_ws, deleted_df)
        style_sheet_as_table(deleted_ws, "DeletedTable")
        summary_ws = wb.create_sheet("サマリー")
        write_summary(summary_ws, aggregate_df, detail_df, deleted_df)
        wb.save(EXCEL_PATH)
        return EXCEL_PATH


def get_excel_bytes() -> bytes | None:
    return EXCEL_PATH.read_bytes() if EXCEL_PATH.exists() else None


def apply_filters(df: pd.DataFrame, department: str, target_year: str, period: str) -> pd.DataFrame:
    result = df.copy()
    if department:
        result = result[result["部署"].astype(str).str.contains(department, case=False, na=False)]
    if target_year:
        result = result[result["年"].astype(str) == target_year]
    if period:
        result = result[result["時期"].astype(str) == period]
    return result


def reset_input_widgets() -> None:
    st.session_state["input_editor_version"] = int(st.session_state.get("input_editor_version", 0)) + 1


def page_input() -> None:
    st.subheader("入力")
    st.caption("部署を選択し、10行すべてのコース欄とコメントを入力して登録してください。")
    if "input_editor_version" not in st.session_state:
        st.session_state["input_editor_version"] = 0
    if success_message := st.session_state.pop("entry_success_message", ""):
        st.success(success_message)
    if info_message := st.session_state.pop("entry_info_message", ""):
        st.info(info_message)

    department = st.selectbox(
        "部署",
        [""] + DEPARTMENT_OPTIONS,
        index=0,
        format_func=lambda value: "選択してください" if value == "" else value,
        key="entry_department",
    )

    st.markdown("#### コース別入力テーブル")
    st.caption("年・時期は固定表示です。Tableau / RPA / ビジネスコア / DBエンジニア / プロ は数値で入力してください。")
    version = st.session_state["input_editor_version"]
    edited_df = st.data_editor(
        make_default_input_df(),
        key=f"course_input_editor_{version}",
        num_rows="fixed",
        use_container_width=True,
        hide_index=True,
        disabled=["年", "時期"],
        column_config={
            "年": st.column_config.NumberColumn("年", format="%d"),
            "時期": st.column_config.TextColumn("時期"),
            "Tableau": st.column_config.NumberColumn("Tableau", min_value=0, step=1, format="%d"),
            "RPA": st.column_config.NumberColumn("RPA", min_value=0, step=1, format="%d"),
            "ビジネスコア": st.column_config.NumberColumn("ビジネスコア", min_value=0, step=1, format="%d"),
            "DBエンジニア": st.column_config.NumberColumn("DBエンジニア", min_value=0, step=1, format="%d"),
            "プロ": st.column_config.NumberColumn("プロ", min_value=0, step=1, format="%d"),
        },
    )

    st.caption("合計")
    st.dataframe(make_input_total_df(edited_df), use_container_width=True, hide_index=True)

    comment = st.text_area(
        "コメント",
        key=f"entry_comment_{version}",
        placeholder="登録内容に関するコメントを入力してください",
    )
    submitted = st.button("確定してSQLiteへ登録し、Excelへ保存", type="primary")

    if submitted:
        department = clean_text(department)
        if not department:
            st.error("部署を選択してください。")
            return
        rows, errors = normalize_input_rows(edited_df, comment)
        if errors:
            st.error("未入力または数値ではない項目があります。すべて数値で入力してから登録してください。")
            for error in errors[:20]:
                st.write(f"- {error}")
            if len(errors) > 20:
                st.write(f"- ほか {len(errors) - 20} 件")
            return
        saved_count = insert_records(department, clean_text(comment), rows)
        excel_path = sync_excel_from_sqlite()
        reset_input_widgets()
        st.session_state["entry_success_message"] = f"{saved_count}件を登録しました。Excelにも保存しました: {excel_path}"
        st.session_state["entry_info_message"] = "登録後、入力テーブルとコメント欄をデフォルトに戻しました。"
        st.rerun()

def page_summary() -> None:
    st.subheader("集計テーブル")
    aggregate_df = fetch_records()
    with st.expander("絞り込み", expanded=True):
        col1, col2, col3 = st.columns(3)
        with col1:
            department = st.selectbox("部署", [""] + DEPARTMENT_OPTIONS, key="filter_department")
        with col2:
            year_options = [""] + sorted([str(x) for x in aggregate_df["年"].dropna().unique()], reverse=True) if not aggregate_df.empty else [""]
            target_year = st.selectbox("年", year_options, key="filter_year")
        with col3:
            period = st.selectbox("時期", ["", "上期", "下期"], key="filter_period")
    filtered_df = apply_filters(aggregate_df, department, target_year, period)
    detail_df = fetch_detail_records()
    deleted_df = fetch_deleted_records()
    c1, c2, c3 = st.columns(3)
    c1.metric("表示中の登録件数", len(filtered_df))
    c2.metric("全登録件数", len(aggregate_df))
    c3.metric("削除履歴件数", len(deleted_df))
    st.dataframe(filtered_df, use_container_width=True, hide_index=True)
    st.markdown("#### Excel")
    if st.button("Excelを再作成", type="secondary"):
        path = sync_excel_from_sqlite()
        st.success(f"Excelを再作成しました: {path}")
    excel_bytes = get_excel_bytes()
    if excel_bytes:
        st.download_button("Excelをダウンロード", data=excel_bytes, file_name="course_entries.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.info("まだExcelファイルが作成されていません。入力画面で確定するか、Excelを再作成してください。")


def page_detail() -> None:
    st.subheader("コース別明細")
    st.caption("Excelの「コース別明細」シートと同じ形式です。1コース1行で確認できます。")
    st.dataframe(fetch_detail_records(), use_container_width=True, hide_index=True)


def page_admin() -> None:
    st.subheader("データ管理")
    st.write("保存先")
    st.code(f"SQLite: {DB_PATH}\nExcel:  {EXCEL_PATH}")
    if success_message := st.session_state.pop("delete_success_message", ""):
        st.success(success_message)
    aggregate_df = fetch_records()
    deleted_df = fetch_deleted_records()
    st.markdown("#### 登録済みデータの削除")
    st.caption("削除する行にチェックを入れ、削除コメントを入力して削除してください。削除した行は削除履歴テーブルに保持され、Excelにも出力されます。")
    if aggregate_df.empty:
        st.info("削除対象の登録データがありません。")
    else:
        delete_df = aggregate_df.copy()
        delete_df.insert(0, "削除", False)
        edited_delete_df = st.data_editor(
            delete_df,
            key="delete_rows_editor",
            use_container_width=True,
            hide_index=True,
            disabled=[col for col in delete_df.columns if col != "削除"],
            column_config={"削除": st.column_config.CheckboxColumn("削除", help="削除する行にチェックを入れてください")},
        )
        delete_comment = st.text_area("削除コメント（必須）", key="delete_comment", placeholder="削除理由を入力してください")
        if st.button("チェックした行を削除", type="primary"):
            selected_ids = edited_delete_df.loc[edited_delete_df["削除"].fillna(False), "ID"].astype(int).tolist()
            if not selected_ids:
                st.warning("削除する行にチェックを入れてください。")
                return
            if not clean_text(delete_comment):
                st.error("削除コメントを入力してください。")
                return
            deleted_count = delete_records(selected_ids, clean_text(delete_comment))
            sync_excel_from_sqlite()
            st.session_state["delete_success_message"] = f"{deleted_count}件を削除し、削除履歴に保存しました。"
            st.rerun()
    st.markdown("#### 削除履歴")
    st.dataframe(deleted_df, use_container_width=True, hide_index=True)
    with st.expander("危険操作: 全データ初期化"):
        st.warning("検証環境用です。登録中データ・削除履歴・監査ログをすべて削除します。")
        confirm = st.text_input("初期化する場合は DELETE と入力")
        if st.button("SQLiteとExcelを初期化", type="secondary", disabled=(confirm != "DELETE")):
            with get_conn() as conn:
                conn.execute("BEGIN IMMEDIATE")
                conn.execute("DELETE FROM audit_logs")
                conn.execute("DELETE FROM deleted_course_records")
                conn.execute("DELETE FROM course_records")
                conn.commit()
            sync_excel_from_sqlite()
            st.success("全データを削除し、Excelを初期化しました。")


def main() -> None:
    st.set_page_config(page_title=APP_TITLE, page_icon="📊", layout="wide", initial_sidebar_state="collapsed")
    init_db()
    st.title(APP_TITLE)
    st.caption("Streamlit + SQLite + Excel保存版")
    menu = st.sidebar.radio("メニュー", ["入力", "集計テーブル", "コース別明細", "データ管理"], index=0)
    if menu == "入力":
        page_input()
    elif menu == "集計テーブル":
        page_summary()
    elif menu == "コース別明細":
        page_detail()
    else:
        page_admin()


if __name__ == "__main__":
    main()
